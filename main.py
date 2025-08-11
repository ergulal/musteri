import sqlite3
import openpyxl
import os

DB_FILE = "musteri_kayit.db"

def db_baglanti():
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""CREATE TABLE IF NOT EXISTS musteriler (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ad_soyad TEXT NOT NULL,
        telefon TEXT,
        email TEXT,
        adres TEXT
    )""")
    return conn

def musteri_ekle(conn):
    ad = input("Ad Soyad: ")
    tel = input("Telefon: ")
    email = input("E-posta: ")
    adres = input("Adres: ")
    conn.execute("INSERT INTO musteriler (ad_soyad, telefon, email, adres) VALUES (?, ?, ?, ?)",
                 (ad, tel, email, adres))
    conn.commit()
    print("Müşteri eklendi.")

def musterileri_listele(conn):
    for row in conn.execute("SELECT * FROM musteriler"):
        print(row)

def musteri_ara(conn):
    kelime = input("Arama: ")
    for row in conn.execute("""SELECT * FROM musteriler 
                                WHERE ad_soyad LIKE ? OR telefon LIKE ? OR email LIKE ? OR adres LIKE ?""",
                            (f"%{kelime}%",)*4):
        print(row)

def musteri_sil(conn):
    id_ = input("Silinecek ID: ")
    conn.execute("DELETE FROM musteriler WHERE id=?", (id_,))
    conn.commit()
    print("Müşteri silindi.")

def musteri_guncelle(conn):
    id_ = input("Güncellenecek ID: ")
    ad = input("Yeni Ad Soyad: ")
    tel = input("Yeni Telefon: ")
    email = input("Yeni E-posta: ")
    adres = input("Yeni Adres: ")
    conn.execute("""UPDATE musteriler SET ad_soyad=?, telefon=?, email=?, adres=? WHERE id=?""",
                 (ad, tel, email, adres, id_))
    conn.commit()
    print("Müşteri güncellendi.")

def excel_aktar(conn):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Müşteriler"
    ws.append(["ID", "Ad Soyad", "Telefon", "E-posta", "Adres"])
    for row in conn.execute("SELECT * FROM musteriler"):
        ws.append(row)
    dosya_adi = "musteriler.xlsx"
    wb.save(dosya_adi)
    print(f"Excel'e aktarıldı: {dosya_adi}")

def excel_ice_aktar(conn):
    dosya_adi = input("İçe aktarılacak Excel dosya adı: ")
    if not os.path.exists(dosya_adi):
        print("Dosya bulunamadı.")
        return
    wb = openpyxl.load_workbook(dosya_adi)
    ws = wb.active
    satir_sayisi = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # row: (ID, ad_soyad, telefon, email, adres)
        if row[1]:  # ad_soyad boş değilse
            conn.execute("INSERT INTO musteriler (ad_soyad, telefon, email, adres) VALUES (?, ?, ?, ?)",
                         (row[1], row[2], row[3], row[4]))
            satir_sayisi += 1
    conn.commit()
    print(f"{satir_sayisi} müşteri içe aktarıldı.")

def menu():
    conn = db_baglanti()
    while True:
        print("""
1) Müşteri Ekle
2) Listele
3) Ara
4) Sil
5) Güncelle
6) Excel'e Aktar
7) Excel'den İçe Aktar
8) Çıkış
""")
        secim = input("Seçim: ")
        if secim == "1": musteri_ekle(conn)
        elif secim == "2": musterileri_listele(conn)
        elif secim == "3": musteri_ara(conn)
        elif secim == "4": musteri_sil(conn)
        elif secim == "5": musteri_guncelle(conn)
        elif secim == "6": excel_aktar(conn)
        elif secim == "7": excel_ice_aktar(conn)
        elif secim == "8": break
        else: print("Hatalı seçim.")

if __name__ == "__main__":
    menu()
